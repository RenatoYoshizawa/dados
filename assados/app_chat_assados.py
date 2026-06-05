from datetime import datetime
from pathlib import Path
from difflib import SequenceMatcher
import re
import unicodedata

import streamlit.components.v1 as components

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURAÇÕES
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

ARQUIVO_CONFIG = BASE_DIR / "config_assados.xlsx"
PASTA_PEDIDOS = BASE_DIR / "pedidos"
ARQUIVO_CARDAPIO_IMG = BASE_DIR / "jmd.png"

st.set_page_config(
    page_title="Chatbot JMD Assados",
    page_icon="🍗",
    layout="wide",
)


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def normalizar_texto(valor):
    if valor is None:
        return ""
    texto = str(valor).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = texto.lower()
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def chave_coluna(valor):
    texto = normalizar_texto(valor)
    texto = re.sub(r"[^a-z0-9]+", "_", texto).strip("_")
    return texto


def texto(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def dinheiro(valor):
    if valor is None or texto(valor) == "":
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    valor = str(valor).replace("R$", "").replace(" ", "").strip()

    if "," in valor:
        valor = valor.replace(".", "").replace(",", ".")

    try:
        return float(valor)
    except ValueError:
        return 0.0


def moeda(valor):
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def sim(valor, padrao=True):
    if valor is None or texto(valor) == "":
        return padrao

    return normalizar_texto(valor) in [
        "sim", "s", "true", "verdadeiro", "1", "ativo", "disponivel", "yes", "y"
    ]


def obter(linha, *nomes, padrao=""):
    for nome in nomes:
        chave = chave_coluna(nome)
        if chave in linha and linha[chave] not in [None, ""]:
            return linha[chave]
    return padrao


def contem_alguma(entrada, termos):
    entrada = normalizar_texto(entrada)
    return any(normalizar_texto(t) in entrada for t in termos)


def solicitar_atendente(msg_norm):
    termos = [
        "atendente",
        "falar com atendente",
        "humano",
        "pessoa",
        "falar com alguem",
        "falar com alguém",
        "suporte",
        "ajuda",
        "preciso de ajuda",
    ]
    return contem_alguma(msg_norm, termos)


def pontuar_texto(entrada, alvo):
    entrada = normalizar_texto(entrada)
    alvo = normalizar_texto(alvo)

    if not entrada or not alvo:
        return 0

    if entrada == alvo:
        return 1.0

    if entrada in alvo or alvo in entrada:
        return 0.92

    tokens_entrada = set(entrada.split())
    tokens_alvo = set(alvo.split())

    if tokens_entrada and tokens_alvo:
        intersecao = tokens_entrada.intersection(tokens_alvo)
        score_tokens = len(intersecao) / max(len(tokens_entrada), 1)
    else:
        score_tokens = 0

    score_similaridade = SequenceMatcher(None, entrada, alvo).ratio()

    return max(score_tokens, score_similaridade)


def resolver_por_numero_ou_palavra(msg_norm, mapa, campo=None, tipo_opcao="opção"):
    """
    Aceita número ou palavra-chave.

    Retorna:
    - ("ok", chave, "")
    - ("ambigua", None, mensagem)
    - ("nao_encontrado", None, "")
    """
    msg_norm = normalizar_texto(msg_norm)

    # Padrão principal: número
    if msg_norm in mapa:
        return "ok", msg_norm, ""

    candidatos = []

    for codigo, valor in mapa.items():
        if isinstance(valor, dict) and campo:
            nome = texto(valor.get(campo, ""))
        else:
            nome = texto(valor)

        score = pontuar_texto(msg_norm, nome)

        if score >= 0.70:
            candidatos.append((score, codigo, nome, valor))

    if not candidatos:
        return "nao_encontrado", None, ""

    candidatos.sort(reverse=True, key=lambda x: x[0])
    melhor_score = candidatos[0][0]

    # Se houver opções muito próximas, pergunta qual é a correta
    candidatos_proximos = [
        item for item in candidatos
        if melhor_score - item[0] <= 0.08 or item[0] >= 0.90
    ]

    if len(candidatos_proximos) == 1:
        return "ok", candidatos_proximos[0][1], ""

    linhas = [f"Encontrei mais de uma {tipo_opcao}. Qual é a correta?\n"]

    for _, codigo, nome, valor in candidatos_proximos:
        if isinstance(valor, dict) and "preco" in valor:
            unidade = f" ({valor.get('unidade', '')})" if valor.get("unidade") else ""
            linhas.append(f"{codigo} - {nome}{unidade} - {moeda(valor['preco'])}")
        elif isinstance(valor, dict) and "taxa" in valor:
            linhas.append(f"{codigo} - {nome} - Taxa {moeda(valor['taxa'])}")
        else:
            linhas.append(f"{codigo} - {nome}")

    linhas.append("\nDigite o número da opção correta.")

    return "ambigua", None, "\n".join(linhas)


def interpretar_menu(msg_norm):
    msg_norm = normalizar_texto(msg_norm)

    if msg_norm in ["1", "2", "3", "4"]:
        return msg_norm

    if contem_alguma(msg_norm, ["pedido", "pedir", "comprar", "quero pedir", "fazer pedido"]):
        return "1"

    if contem_alguma(msg_norm, ["cardapio", "cardápio", "menu", "preco", "preço", "valores"]):
        return "2"

    if contem_alguma(msg_norm, ["horario", "horário", "funcionamento", "abre", "fecha"]):
        return "3"

    if solicitar_atendente(msg_norm):
        return "4"

    return ""


def interpretar_sim_nao(msg_norm):
    msg_norm = normalizar_texto(msg_norm)

    if msg_norm in ["1", "sim", "s", "confirmo", "confirmar", "ok", "pode", "isso", "correto"]:
        return "sim"

    if msg_norm in ["2", "nao", "não", "n", "cancelar", "cancela", "errado"]:
        return "nao"

    return ""


def interpretar_tipo_entrega(msg_norm):
    msg_norm = normalizar_texto(msg_norm)

    if msg_norm == "1" or contem_alguma(msg_norm, ["retirada", "retirar", "buscar", "busco", "loja"]):
        return "retirada"

    if msg_norm == "2" or contem_alguma(msg_norm, ["entrega", "entregar", "delivery", "endereco", "endereço"]):
        return "entrega"

    return ""


def interpretar_mais_itens(msg_norm):
    msg_norm = normalizar_texto(msg_norm)

    if msg_norm == "1" or contem_alguma(msg_norm, ["mesma", "mesma categoria", "outro item", "mais um"]):
        return "mesma_categoria"

    if msg_norm == "2" or contem_alguma(msg_norm, ["outra categoria", "categoria", "outro tipo"]):
        return "outra_categoria"

    if msg_norm == "3" or contem_alguma(msg_norm, ["resumo", "ver pedido", "parcial"]):
        return "resumo"

    if msg_norm == "4" or contem_alguma(msg_norm, ["finalizar", "fechar", "concluir", "nao", "não", "só isso", "so isso"]):
        return "finalizar"

    return ""


@st.cache_data(show_spinner=False)
def carregar_aba_excel(caminho, nome_aba):
    caminho = Path(caminho)

    if not caminho.exists():
        return []

    wb = load_workbook(caminho, data_only=True)

    if nome_aba not in wb.sheetnames:
        return []

    ws = wb[nome_aba]
    cabecalhos = [chave_coluna(c.value) for c in ws[1]]
    linhas = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(valor not in [None, ""] for valor in row):
            continue

        item = {}
        for i, valor in enumerate(row):
            if i < len(cabecalhos) and cabecalhos[i]:
                item[cabecalhos[i]] = valor

        linhas.append(item)

    return linhas


# ============================================================
# LEITURA DO CONFIG_ASSADOS.XLSX
# ============================================================

@st.cache_data(show_spinner=False)
def carregar_cardapio():
    linhas = carregar_aba_excel(ARQUIVO_CONFIG, "Cardapio")
    cardapio = []

    for i, linha in enumerate(linhas, start=1):
        codigo = texto(obter(linha, "Código", "Codigo", padrao=i)).upper()
        produto = texto(obter(linha, "Produto", "Nome", "Item"))

        if not produto:
            continue

        categoria = texto(obter(linha, "Categoria", padrao="Geral"))
        unidade = texto(obter(linha, "Unidade", "Tamanho", padrao="Unidade"))
        preco = dinheiro(obter(linha, "Preço", "Preco", "Valor"))
        disponivel = sim(obter(linha, "Disponível", "Disponivel", padrao="SIM"))
        observacao = texto(obter(linha, "Observação", "Observacao", "Descrição", "Descricao"))

        if disponivel:
            cardapio.append({
                "codigo_original": codigo,
                "categoria": categoria,
                "produto": produto,
                "unidade": unidade,
                "preco": preco,
                "observacao": observacao,
                "tipo": "produto",
            })

    combos = carregar_aba_excel(ARQUIVO_CONFIG, "Combos_Promocoes")

    for i, linha in enumerate(combos, start=1):
        codigo = texto(obter(linha, "Código", "Codigo", padrao=f"C{i}")).upper()
        nome_combo = texto(obter(linha, "Nome do combo", "Promoção", "Promocao", "Combo", "Nome"))

        if not nome_combo:
            continue

        disponivel = sim(obter(linha, "Disponível", "Disponivel", padrao="SIM"))
        if not disponivel:
            continue

        itens = texto(obter(linha, "Itens sugeridos", "Itens", "Descrição", "Descricao"))
        preco = dinheiro(obter(linha, "Preço", "Preco", "Valor"))

        cardapio.append({
            "codigo_original": codigo,
            "categoria": "Combos e promoções",
            "produto": nome_combo,
            "unidade": "Combo",
            "preco": preco,
            "observacao": itens,
            "tipo": "combo",
        })

    return cardapio


@st.cache_data(show_spinner=False)
def carregar_locais_entrega():
    linhas = carregar_aba_excel(ARQUIVO_CONFIG, "Locais_Entrega")
    locais = []

    for i, linha in enumerate(linhas, start=1):
        codigo = texto(obter(linha, "Código", "Codigo", padrao=i)).upper()
        bairro = texto(obter(
            linha,
            "Bairro",
            "Bairro/Região",
            "Bairro/Regiao",
            "Local",
            "Região",
            "Regiao",
        ))

        if not bairro:
            continue

        taxa = dinheiro(obter(linha, "Taxa", "Taxa de entrega", "Valor"))
        tempo_estimado = texto(obter(linha, "Tempo estimado", "Tempo", "Prazo"))
        atende = sim(obter(linha, "Atende?", "Atende", "Disponível", "Disponivel", padrao="SIM"))
        bairro_norm = normalizar_texto(bairro)

        if atende and "retirada" not in bairro_norm:
            locais.append({
                "codigo_original": codigo,
                "bairro": bairro,
                "taxa": taxa,
                "tempo_estimado": tempo_estimado,
            })

    return locais


@st.cache_data(show_spinner=False)
def carregar_formas_pagamento():
    linhas = carregar_aba_excel(ARQUIVO_CONFIG, "Formas_Pagamento")
    pagamentos = []

    for i, linha in enumerate(linhas, start=1):
        codigo = texto(obter(linha, "Código", "Codigo", padrao=i)).upper()
        forma = texto(obter(linha, "Forma", "Forma de pagamento", "Pagamento", "Nome"))

        if not forma:
            continue

        disponivel = sim(obter(linha, "Disponível", "Disponivel", padrao="SIM"))
        observacao = texto(obter(linha, "Observação", "Observacao"))
        perguntar_troco = sim(
            obter(linha, "Perguntar troco?", "Perguntar troco", padrao="NÃO"),
            padrao=False,
        )

        if disponivel:
            pagamentos.append({
                "codigo_original": codigo,
                "forma": forma,
                "observacao": observacao,
                "perguntar_troco": perguntar_troco,
            })

    if not pagamentos:
        pagamentos = [
            {"codigo_original": "1", "forma": "Pix", "observacao": "", "perguntar_troco": False},
            {"codigo_original": "2", "forma": "Cartão", "observacao": "", "perguntar_troco": False},
            {"codigo_original": "3", "forma": "Dinheiro", "observacao": "", "perguntar_troco": True},
        ]

    return pagamentos


@st.cache_data(show_spinner=False)
def carregar_configuracoes():
    linhas = carregar_aba_excel(ARQUIVO_CONFIG, "Configuracoes")
    configs = {}

    for linha in linhas:
        campo = normalizar_texto(obter(linha, "Campo", "Chave", "Configuração", "Configuracao"))
        valor = texto(obter(linha, "Valor", "Conteúdo", "Conteudo"))

        if campo:
            configs[campo] = valor

    return configs


@st.cache_data(show_spinner=False)
def carregar_horarios():
    linhas = carregar_aba_excel(ARQUIVO_CONFIG, "Horarios")
    horarios = []

    for linha in linhas:
        dia = texto(obter(linha, "Dia", "Dia da semana", padrao=""))
        abre = texto(obter(linha, "Abre", "Abertura", padrao=""))
        fecha = texto(obter(linha, "Fecha", "Fechamento", padrao=""))
        aceita = texto(obter(linha, "Aceita pedido?", "Aceita pedido", padrao="SIM"))
        observacao = texto(obter(linha, "Observação", "Observacao", padrao=""))

        if dia:
            horarios.append({
                "dia": dia,
                "abre": abre,
                "fecha": fecha,
                "aceita": aceita,
                "observacao": observacao,
            })

    return horarios


def nome_loja():
    configs = carregar_configuracoes()
    return configs.get("nome_loja") or configs.get("nome da loja") or "JMD Assados"


def chave_pix():
    configs = carregar_configuracoes()
    return configs.get("chave_pix") or configs.get("pix") or "PREENCHER"


def telefone_atendente():
    configs = carregar_configuracoes()
    return configs.get("telefone_atendente") or configs.get("telefone") or "PREENCHER"


# ============================================================
# PLANILHA DIÁRIA DE PEDIDOS
# ============================================================

def criar_ou_abrir_planilha_pedidos():
    Path(PASTA_PEDIDOS).mkdir(exist_ok=True)

    data_arquivo = datetime.now().strftime("%d-%m-%Y")
    caminho = Path(PASTA_PEDIDOS) / f"pedidos_{data_arquivo}.xlsx"

    if caminho.exists():
        wb = load_workbook(caminho)
        return wb, caminho

    wb = Workbook()

    ws = wb.active
    ws.title = "Pedidos"
    ws.append([
        "Nº Pedido",
        "Data/Hora",
        "Cliente",
        "Telefone",
        "Tipo",
        "Bairro",
        "Endereço",
        "Pagamento",
        "Troco",
        "Taxa Entrega",
        "Total",
        "Status",
    ])

    ws_itens = wb.create_sheet("Itens_Pedido")
    ws_itens.append([
        "Nº Pedido",
        "Produto",
        "Categoria",
        "Unidade",
        "Quantidade",
        "Valor Unitário",
        "Subtotal",
    ])

    ws_resumo = wb.create_sheet("Resumo_Dia")
    ws_resumo.append(["Indicador", "Valor"])
    ws_resumo.append(["Total de pedidos", "=COUNTA(Pedidos!A:A)-1"])
    ws_resumo.append(["Total vendido", "=SUM(Pedidos!K:K)"])
    ws_resumo.append(["Total taxas de entrega", "=SUM(Pedidos!J:J)"])
    ws_resumo.append(["Pedidos entrega", '=COUNTIF(Pedidos!E:E,"Entrega")'])
    ws_resumo.append(["Pedidos retirada", '=COUNTIF(Pedidos!E:E,"Retirada")'])

    formatar_planilha(wb)
    wb.save(caminho)

    return wb, caminho


def formatar_planilha(wb):
    cor_cabecalho = "7F1D1D"

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=cor_cabecalho)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        for coluna in range(1, ws.max_column + 1):
            letra = get_column_letter(coluna)
            ws.column_dimensions[letra].width = 18

        if ws.title == "Pedidos":
            ws.column_dimensions["C"].width = 25
            ws.column_dimensions["G"].width = 40

        if ws.title == "Itens_Pedido":
            ws.column_dimensions["B"].width = 35


def proximo_numero_pedido(ws):
    maior = 0

    for linha in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        valor = linha[0]
        if valor is None:
            continue

        try:
            maior = max(maior, int(str(valor)))
        except ValueError:
            pass

    return f"{maior + 1:04d}"


def salvar_pedido_excel():
    wb, caminho = criar_ou_abrir_planilha_pedidos()

    ws = wb["Pedidos"]
    ws_itens = wb["Itens_Pedido"]

    numero = proximo_numero_pedido(ws)
    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    ws.append([
        numero,
        agora,
        st.session_state.nome_cliente,
        st.session_state.telefone_cliente,
        st.session_state.tipo_entrega,
        st.session_state.bairro,
        st.session_state.endereco,
        st.session_state.pagamento,
        st.session_state.troco,
        st.session_state.taxa_entrega,
        total_pedido(),
        "Confirmado",
    ])

    for item in st.session_state.pedido:
        ws_itens.append([
            numero,
            item["produto"],
            item["categoria"],
            item["unidade"],
            item["quantidade"],
            item["preco"],
            item["subtotal"],
        ])

    for cell in ws["J"][1:]:
        cell.number_format = 'R$ #,##0.00'

    for cell in ws["K"][1:]:
        cell.number_format = 'R$ #,##0.00'

    for cell in ws_itens["F"][1:]:
        cell.number_format = 'R$ #,##0.00'

    for cell in ws_itens["G"][1:]:
        cell.number_format = 'R$ #,##0.00'

    wb.save(caminho)
    return numero, caminho


# ============================================================
# ESTADO DO CHAT
# ============================================================

def estado_padrao():
    return {
        "etapa": "inicio",
        "messages": [
            {
                "role": "assistant",
                "content": (
                    f"Olá! Bem-vindo(a) à {nome_loja()}.\n\n"
                    "Envie **oi** para iniciar o atendimento."
                ),
            }
        ],
        "pedido": [],
        "categoria_atual": "",
        "produto_atual": None,
        "tipo_entrega": "",
        "bairro": "",
        "endereco": "",
        "nome_cliente": "",
        "telefone_cliente": "",
        "pagamento": "",
        "troco": "",
        "taxa_entrega": 0.0,
        "numero_pedido": "",
        "arquivo_pedido": "",
        "mostrar_imagem_cardapio": False,
        "mapa_categorias": {},
        "mapa_produtos": {},
        "mapa_locais": {},
        "mapa_pagamentos": {},
    }


def inicializar_estado():
    if "etapa" not in st.session_state:
        for chave, valor in estado_padrao().items():
            st.session_state[chave] = valor


def reiniciar_atendimento(mensagem_inicial=True):
    novo = estado_padrao()
    for chave, valor in novo.items():
        st.session_state[chave] = valor

    if not mensagem_inicial:
        st.session_state.messages = []


def add_user(msg):
    st.session_state.messages.append({"role": "user", "content": msg})


def add_bot(msg):
    st.session_state.messages.append({"role": "assistant", "content": msg})


# ============================================================
# TEXTO DAS ETAPAS
# ============================================================

def listar_categorias():
    categorias = []

    for item in carregar_cardapio():
        categoria = item["categoria"]
        if categoria not in categorias:
            categorias.append(categoria)

    return categorias


def produtos_por_categoria(categoria):
    return [item for item in carregar_cardapio() if item["categoria"] == categoria]


def categoria_exige_maioridade(categoria):
    categoria_norm = normalizar_texto(categoria)
    return "alcool" in categoria_norm or "alcoo" in categoria_norm


def texto_menu():
    return (
        "Escolha uma opção:\n\n"
        "1 - Fazer pedido\n"
        "2 - Ver cardápio\n"
        "3 - Consultar horário de funcionamento\n"
        "4 - Falar com atendente\n\n"
        "Digite o número da opção desejada."
    )


def texto_categorias():
    categorias = listar_categorias()
    st.session_state.mapa_categorias = {}

    linhas = ["Escolha uma categoria:\n"]

    for i, categoria in enumerate(categorias, start=1):
        codigo = str(i)
        st.session_state.mapa_categorias[codigo] = categoria
        linhas.append(f"{codigo} - {categoria}")

    linhas.append("\n0 - Voltar ao menu")
    return "\n".join(linhas)


def texto_produtos(categoria):
    produtos = produtos_por_categoria(categoria)
    st.session_state.mapa_produtos = {}

    linhas = [f"{categoria.upper()}:\n"]

    for i, item in enumerate(produtos, start=1):
        codigo = str(i)
        st.session_state.mapa_produtos[codigo] = item
        unidade = f" ({item['unidade']})" if item.get("unidade") else ""
        obs = f" - {item['observacao']}" if item.get("observacao") else ""
        linhas.append(f"{codigo} - {item['produto']}{unidade} - {moeda(item['preco'])}{obs}")

    linhas.append("\nDigite o número do item desejado.")
    linhas.append("0 - Voltar às categorias")
    return "\n".join(linhas)


def texto_horarios():
    horarios = carregar_horarios()

    if not horarios:
        return "Horário de funcionamento ainda não configurado."

    linhas = ["Horário de funcionamento:\n"]

    for h in horarios:
        obs = f" | {h['observacao']}" if h["observacao"] else ""
        linhas.append(
            f"{h['dia']}: {h['abre']} às {h['fecha']} | Aceita pedido: {h['aceita']}{obs}"
        )

    linhas.append("\nDigite **1** para fazer pedido ou **0** para voltar ao menu.")
    return "\n".join(linhas)


def texto_locais():
    locais = carregar_locais_entrega()
    st.session_state.mapa_locais = {}

    if not locais:
        return "Não há bairros cadastrados. Digite o endereço completo para entrega."

    linhas = ["Selecione o bairro/região para entrega:\n"]

    for i, local in enumerate(locais, start=1):
        codigo = str(i)
        st.session_state.mapa_locais[codigo] = local
        tempo = f" | {local['tempo_estimado']}" if local["tempo_estimado"] else ""
        linhas.append(f"{codigo} - {local['bairro']} - Taxa {moeda(local['taxa'])}{tempo}")

    linhas.append("\nDigite o número do bairro/região.")
    return "\n".join(linhas)


def texto_pagamentos():
    pagamentos = carregar_formas_pagamento()
    st.session_state.mapa_pagamentos = {}

    linhas = ["Escolha a forma de pagamento:\n"]

    for i, pg in enumerate(pagamentos, start=1):
        codigo = str(i)
        st.session_state.mapa_pagamentos[codigo] = pg
        obs = f" - {pg['observacao']}" if pg["observacao"] else ""
        linhas.append(f"{codigo} - {pg['forma']}{obs}")

    linhas.append("\nDigite o número da forma de pagamento.")
    return "\n".join(linhas)


def total_itens():
    return sum(item["subtotal"] for item in st.session_state.pedido)


def total_pedido():
    return total_itens() + float(st.session_state.taxa_entrega or 0)


def texto_resumo_parcial():
    if not st.session_state.pedido:
        return "Seu pedido ainda não possui itens."

    linhas = ["Resumo atual do pedido:\n"]

    for item in st.session_state.pedido:
        linhas.append(f"{item['quantidade']}x {item['produto']} - {moeda(item['subtotal'])}")

    linhas.append(f"\nSubtotal dos itens: {moeda(total_itens())}")

    if st.session_state.taxa_entrega:
        linhas.append(f"Taxa de entrega: {moeda(st.session_state.taxa_entrega)}")

    linhas.append(f"Total parcial: {moeda(total_pedido())}")
    return "\n".join(linhas)


def texto_pergunta_mais_itens():
    return (
        "Deseja adicionar mais algum item?\n\n"
        "1 - Sim, da mesma categoria\n"
        "2 - Sim, escolher outra categoria\n"
        "3 - Ver resumo atual\n"
        "4 - Não, finalizar pedido"
    )


def texto_resumo_final():
    linhas = ["Confira o resumo do pedido:\n"]

    for item in st.session_state.pedido:
        linhas.append(
            f"{item['quantidade']}x {item['produto']} - "
            f"{moeda(item['preco'])} cada | Subtotal: {moeda(item['subtotal'])}"
        )

    linhas.append(f"\nSubtotal: {moeda(total_itens())}")

    if st.session_state.taxa_entrega:
        linhas.append(f"Taxa de entrega: {moeda(st.session_state.taxa_entrega)}")

    linhas.append(f"Total: {moeda(total_pedido())}")
    linhas.append(f"\nCliente: {st.session_state.nome_cliente}")
    linhas.append(f"Telefone: {st.session_state.telefone_cliente}")
    linhas.append(f"Tipo: {st.session_state.tipo_entrega}")

    if st.session_state.tipo_entrega == "Entrega":
        linhas.append(f"Bairro: {st.session_state.bairro}")
        linhas.append(f"Endereço: {st.session_state.endereco}")

    linhas.append(f"Pagamento: {st.session_state.pagamento}")

    if st.session_state.troco:
        linhas.append(f"Troco: {st.session_state.troco}")

    linhas.append("\nConfirmar pedido?\n1 - Sim\n2 - Cancelar")

    return "\n".join(linhas)


def dataframe_itens():
    if not st.session_state.pedido:
        return pd.DataFrame(columns=["Produto", "Categoria", "Quantidade", "Valor unitário", "Subtotal"])

    return pd.DataFrame([
        {
            "Produto": item["produto"],
            "Categoria": item["categoria"],
            "Quantidade": item["quantidade"],
            "Valor unitário": moeda(item["preco"]),
            "Subtotal": moeda(item["subtotal"]),
        }
        for item in st.session_state.pedido
    ])


# ============================================================
# PROCESSAMENTO DO CHAT
# ============================================================

def processar_mensagem(msg):
    msg_original = msg.strip()
    msg_norm = normalizar_texto(msg_original)

    # Atendente em qualquer etapa do atendimento
    if solicitar_atendente(msg_norm):
        st.session_state.etapa = "atendente"
        st.session_state.mostrar_imagem_cardapio = False
        return (
            "Encaminhando para o atendente.\n\n"
            "Um atendente dará continuidade ao atendimento."
        )

    if msg_norm in ["0", "menu", "inicio", "início", "reiniciar"]:
        reiniciar_atendimento(mensagem_inicial=False)
        st.session_state.etapa = "menu"
        return "Atendimento reiniciado.\n\n" + texto_menu()

    etapa = st.session_state.etapa

    if etapa == "inicio":
        st.session_state.etapa = "menu"
        return texto_menu()

    if etapa == "menu":
        opcao_menu = interpretar_menu(msg_norm)

        if opcao_menu == "1":
            st.session_state.etapa = "categoria"
            st.session_state.mostrar_imagem_cardapio = False
            return texto_categorias()

        if opcao_menu == "2":
            st.session_state.mostrar_imagem_cardapio = True
            return (
                "Claro! Segue a foto do cardápio.\n\n"
                "Para fazer um pedido, digite **1**.\n"
                "Para voltar ao menu, digite **0**."
            )

        if opcao_menu == "3":
            st.session_state.mostrar_imagem_cardapio = False
            return texto_horarios()

        if opcao_menu == "4":
            st.session_state.mostrar_imagem_cardapio = False
            st.session_state.etapa = "atendente"
            return (
                "Encaminhando para o atendente.\n\n"
                "Um atendente dará continuidade ao atendimento."
            )

        return "Opção inválida.\n\n" + texto_menu()

    if etapa == "categoria":
        mapa = st.session_state.mapa_categorias or {}

        if not mapa:
            texto_categorias()
            mapa = st.session_state.mapa_categorias

        status, chave, mensagem_ambigua = resolver_por_numero_ou_palavra(
            msg_norm,
            mapa,
            campo=None,
            tipo_opcao="categoria"
        )

        if status == "ambigua":
            return mensagem_ambigua

        if status != "ok":
            return "Categoria inválida.\n\n" + texto_categorias()

        categoria = mapa[chave]
        st.session_state.categoria_atual = categoria

        if categoria_exige_maioridade(categoria):
            st.session_state.etapa = "maioridade"
            return (
                "Esta categoria possui bebidas alcoólicas.\n"
                "A venda é permitida somente para maiores de 18 anos.\n\n"
                "Confirma que o cliente é maior de 18 anos?\n"
                "1 - Sim\n"
                "2 - Não"
            )

        st.session_state.etapa = "produto"
        return texto_produtos(categoria)

    if etapa == "maioridade":
        resposta_sim_nao = interpretar_sim_nao(msg_norm)

        if resposta_sim_nao == "sim":
            st.session_state.etapa = "produto"
            return texto_produtos(st.session_state.categoria_atual)

        if resposta_sim_nao == "nao":
            st.session_state.etapa = "categoria"
            return "Categoria não liberada.\n\n" + texto_categorias()

        return "Opção inválida.\n\nConfirma que o cliente é maior de 18 anos?\n1 - Sim\n2 - Não"

    if etapa == "produto":
        if msg_norm == "0":
            st.session_state.etapa = "categoria"
            return texto_categorias()

        mapa = st.session_state.mapa_produtos or {}

        if not mapa:
            texto_produtos(st.session_state.categoria_atual)
            mapa = st.session_state.mapa_produtos

        status, chave, mensagem_ambigua = resolver_por_numero_ou_palavra(
            msg_norm,
            mapa,
            campo="produto",
            tipo_opcao="opção de produto"
        )

        if status == "ambigua":
            return mensagem_ambigua

        if status != "ok":
            return "Produto inválido.\n\n" + texto_produtos(st.session_state.categoria_atual)

        produto = mapa[chave]
        st.session_state.produto_atual = produto
        st.session_state.etapa = "quantidade"

        return f"Você escolheu: {produto['produto']}.\nDigite a quantidade desejada."

    if etapa == "quantidade":
        if not msg_norm.isdigit() or int(msg_norm) <= 0:
            return "Digite uma quantidade válida. Exemplo: 1"

        quantidade = int(msg_norm)
        produto = st.session_state.produto_atual
        subtotal = produto["preco"] * quantidade

        st.session_state.pedido.append({
            "produto": produto["produto"],
            "categoria": produto["categoria"],
            "unidade": produto["unidade"],
            "preco": produto["preco"],
            "quantidade": quantidade,
            "subtotal": subtotal,
        })

        st.session_state.produto_atual = None
        st.session_state.etapa = "mais_itens"

        return "Item adicionado ao pedido.\n\n" + texto_pergunta_mais_itens()

    if etapa == "mais_itens":
        acao = interpretar_mais_itens(msg_norm)

        if acao == "mesma_categoria":
            st.session_state.etapa = "produto"
            return texto_produtos(st.session_state.categoria_atual)

        if acao == "outra_categoria":
            st.session_state.etapa = "categoria"
            return texto_categorias()

        if acao == "resumo":
            return texto_resumo_parcial() + "\n\n" + texto_pergunta_mais_itens()

        if acao == "finalizar":
            if not st.session_state.pedido:
                st.session_state.etapa = "categoria"
                return "Pedido sem itens.\n\n" + texto_categorias()

            st.session_state.etapa = "tipo_entrega"
            return "Como deseja receber o pedido?\n\n1 - Retirada na loja\n2 - Entrega"

        return "Opção inválida.\n\n" + texto_pergunta_mais_itens()

    if etapa == "tipo_entrega":
        tipo = interpretar_tipo_entrega(msg_norm)

        if tipo == "retirada":
            st.session_state.tipo_entrega = "Retirada"
            st.session_state.bairro = ""
            st.session_state.endereco = ""
            st.session_state.taxa_entrega = 0.0
            st.session_state.etapa = "nome_cliente"
            return "Informe o nome para identificação do pedido."

        if tipo == "entrega":
            st.session_state.tipo_entrega = "Entrega"
            st.session_state.etapa = "bairro"
            return texto_locais()

        return "Opção inválida.\n\nComo deseja receber o pedido?\n1 - Retirada na loja\n2 - Entrega"

    if etapa == "bairro":
        mapa = st.session_state.mapa_locais or {}

        if not mapa:
            texto_locais()
            mapa = st.session_state.mapa_locais

        if not mapa:
            st.session_state.etapa = "endereco"
            return "Digite o endereço completo para entrega."

        status, chave, mensagem_ambigua = resolver_por_numero_ou_palavra(
            msg_norm,
            mapa,
            campo="bairro",
            tipo_opcao="bairro/região"
        )

        if status == "ambigua":
            return mensagem_ambigua

        if status != "ok":
            return "Bairro/região não localizado.\n\n" + texto_locais()

        local = mapa[chave]
        st.session_state.bairro = local["bairro"]
        st.session_state.taxa_entrega = local["taxa"]
        st.session_state.etapa = "endereco"

        return f"Entrega selecionada para {local['bairro']}.\nDigite o endereço completo."

    if etapa == "endereco":
        if len(msg_original) < 5:
            return "Informe o endereço completo, com rua, número e complemento, se houver."

        st.session_state.endereco = msg_original
        st.session_state.etapa = "nome_cliente"
        return "Informe o nome para identificação do pedido."

    if etapa == "nome_cliente":
        if len(msg_original) < 2:
            return "Informe o nome do cliente."

        st.session_state.nome_cliente = msg_original
        st.session_state.etapa = "telefone_cliente"
        return "Informe o telefone/WhatsApp do cliente."

    if etapa == "telefone_cliente":
        if len(msg_original) < 6:
            return "Informe um telefone válido ou digite 'não informado'."

        st.session_state.telefone_cliente = msg_original
        st.session_state.etapa = "pagamento"
        return texto_pagamentos()

    if etapa == "pagamento":
        mapa = st.session_state.mapa_pagamentos or {}

        if not mapa:
            texto_pagamentos()
            mapa = st.session_state.mapa_pagamentos

        status, chave, mensagem_ambigua = resolver_por_numero_ou_palavra(
            msg_norm,
            mapa,
            campo="forma",
            tipo_opcao="forma de pagamento"
        )

        if status == "ambigua":
            return mensagem_ambigua

        if status != "ok":
            return "Forma de pagamento inválida.\n\n" + texto_pagamentos()

        pagamento = mapa[chave]
        st.session_state.pagamento = pagamento["forma"]

        if pagamento["perguntar_troco"] or "dinheiro" in normalizar_texto(pagamento["forma"]):
            st.session_state.etapa = "troco"
            return "Precisa de troco? Se sim, informe para quanto. Se não, digite NÃO."

        st.session_state.etapa = "confirmacao"
        resposta = texto_resumo_final()

        if "pix" in normalizar_texto(pagamento["forma"]):
            resposta += f"\n\nChave Pix: {chave_pix()}"

        return resposta

    if etapa == "troco":
        st.session_state.troco = msg_original
        st.session_state.etapa = "confirmacao"
        return texto_resumo_final()

    if etapa == "confirmacao":
        resposta_sim_nao = interpretar_sim_nao(msg_norm)

        if resposta_sim_nao == "sim":
            numero, caminho = salvar_pedido_excel()
            st.session_state.numero_pedido = numero
            st.session_state.arquivo_pedido = str(caminho)
            st.session_state.etapa = "final"

            return (
                "Pedido confirmado com sucesso!\n\n"
                f"Nº do pedido: {numero}\n"
                f"Total: {moeda(total_pedido())}\n"
                f"Arquivo gerado/atualizado: {caminho}\n\n"
                "Digite **novo** para iniciar outro atendimento."
            )

        if resposta_sim_nao == "nao":
            reiniciar_atendimento(mensagem_inicial=False)
            st.session_state.etapa = "inicio"
            return "Pedido cancelado. Envie **oi** para iniciar novamente."

        return "Opção inválida.\n\nConfirmar pedido?\n1 - Sim\n2 - Cancelar"

    if etapa == "atendente":
        return (
            "Atendimento encaminhado.\n\n"
            "Um atendente dará continuidade ao atendimento."
        )

    if etapa == "final":
        if msg_norm in ["novo", "oi", "ola", "olá"]:
            reiniciar_atendimento(mensagem_inicial=False)
            st.session_state.etapa = "menu"
            return "Novo atendimento iniciado.\n\n" + texto_menu()

        return (
            "Este pedido já foi finalizado.\n\n"
            "Digite **novo** para iniciar outro atendimento ou **0** para reiniciar."
        )

    return "Não entendi sua mensagem. Digite **0** para reiniciar o atendimento."


# ============================================================
# INTERFACE STREAMLIT EM FORMATO DE CHAT
# ============================================================

def aplicar_css():
    st.markdown(
        """
        <style>
        .stApp {
            background: #fff8f0;
        }
        .pedido-card {
            background: #ffffff;
            border: 1px solid #ead7c2;
            border-radius: 16px;
            padding: 16px;
            margin-bottom: 14px;
        }
        .pedido-total {
            color: #166534;
            font-size: 24px;
            font-weight: 800;
        }
        .pedido-numero {
            color: #7F1D1D;
            font-size: 22px;
            font-weight: 800;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def exibir_sidebar():
    with st.sidebar:
        st.header("🧾 Pedido")

        if st.button("Reiniciar atendimento", use_container_width=True):
            reiniciar_atendimento()
            st.rerun()

        if st.button("Recarregar planilha", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

        st.divider()

        if st.session_state.pedido:
            st.dataframe(dataframe_itens(), use_container_width=True, hide_index=True)
        else:
            st.info("Nenhum item adicionado.")

        st.metric("Subtotal", moeda(total_itens()))
        st.metric("Taxa de entrega", moeda(st.session_state.taxa_entrega))
        st.markdown(f'<div class="pedido-total">Total: {moeda(total_pedido())}</div>', unsafe_allow_html=True)

        if st.session_state.numero_pedido:
            st.divider()
            st.markdown(f'<div class="pedido-numero">Pedido nº {st.session_state.numero_pedido}</div>', unsafe_allow_html=True)

            if st.session_state.arquivo_pedido:
                st.caption(st.session_state.arquivo_pedido)
                caminho = Path(st.session_state.arquivo_pedido)

                if caminho.exists() and caminho.is_file():
                    with open(caminho, "rb") as arquivo:
                        st.download_button(
                            "Baixar planilha do dia",
                            data=arquivo,
                            file_name=caminho.name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )


def exibir_resumo_final_na_tela():
    if st.session_state.etapa != "final":
        return

    st.markdown("### Pedido finalizado")

    st.markdown('<div class="pedido-card">', unsafe_allow_html=True)
    st.markdown(f'<div class="pedido-numero">Pedido nº {st.session_state.numero_pedido}</div>', unsafe_allow_html=True)
    st.write(f"**Cliente:** {st.session_state.nome_cliente}")
    st.write(f"**Telefone:** {st.session_state.telefone_cliente}")
    st.write(f"**Tipo:** {st.session_state.tipo_entrega}")

    if st.session_state.tipo_entrega == "Entrega":
        st.write(f"**Bairro:** {st.session_state.bairro}")
        st.write(f"**Endereço:** {st.session_state.endereco}")

    st.write(f"**Pagamento:** {st.session_state.pagamento}")

    if st.session_state.troco:
        st.write(f"**Troco:** {st.session_state.troco}")

    st.dataframe(dataframe_itens(), use_container_width=True, hide_index=True)

    st.write(f"Subtotal: **{moeda(total_itens())}**")
    st.write(f"Taxa de entrega: **{moeda(st.session_state.taxa_entrega)}**")
    st.markdown(f'<div class="pedido-total">Total: {moeda(total_pedido())}</div>', unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)


def placeholder_chat():
    etapa = st.session_state.etapa

    placeholders = {
        "inicio": "Digite oi para iniciar...",
        "menu": "Digite 1, 2, 3 ou 4...",
        "categoria": "Digite o número da categoria ou uma palavra...",
        "maioridade": "Digite 1 para sim ou 2 para não...",
        "produto": "Digite o número do produto ou uma palavra...",
        "quantidade": "Digite a quantidade...",
        "mais_itens": "Digite 1, 2, 3 ou 4...",
        "tipo_entrega": "Digite 1 para retirada ou 2 para entrega...",
        "bairro": "Digite o número ou nome do bairro/região...",
        "endereco": "Digite o endereço completo...",
        "nome_cliente": "Digite o nome do cliente...",
        "telefone_cliente": "Digite o telefone/WhatsApp...",
        "pagamento": "Digite o número ou nome da forma de pagamento...",
        "troco": "Digite o valor para troco ou NÃO...",
        "confirmacao": "Digite 1 para confirmar ou 2 para cancelar...",
        "atendente": "Atendimento encaminhado...",
        "final": "Digite novo para iniciar outro atendimento...",
    }

    return placeholders.get(etapa, "Digite sua mensagem...")


def rolar_e_focar_chat():
    components.html(
        """
        <script>
        function aplicarAjustes() {
            const doc = window.parent.document;

            const inputContainer =
                doc.querySelector('[data-testid="stChatInput"]') ||
                doc.querySelector('[data-testid="stChatInput"] textarea') ||
                doc.querySelector('textarea[placeholder]');

            if (inputContainer) {
                try {
                    inputContainer.scrollIntoView({
                        behavior: "smooth",
                        block: "center",
                        inline: "nearest"
                    });
                } catch (e) {}
            }

            const inputChat =
                doc.querySelector('[data-testid="stChatInput"] textarea') ||
                doc.querySelector('textarea[aria-label="Chat input"]') ||
                doc.querySelector('textarea[placeholder]') ||
                doc.querySelector('textarea');

            if (inputChat) {
                try {
                    inputChat.focus({ preventScroll: true });
                    const tamanho = inputChat.value.length;
                    inputChat.setSelectionRange(tamanho, tamanho);
                } catch (e) {
                    try {
                        inputChat.focus();
                    } catch (e2) {}
                }
            }
        }

        setTimeout(aplicarAjustes, 100);
        setTimeout(aplicarAjustes, 300);
        setTimeout(aplicarAjustes, 700);
        setTimeout(aplicarAjustes, 1200);
        setTimeout(aplicarAjustes, 1800);
        </script>
        """,
        height=1,
    )


def exibir_cardapio_imagem():
    if not st.session_state.get("mostrar_imagem_cardapio"):
        return

    st.markdown("### 📷 Cardápio")

    if ARQUIVO_CARDAPIO_IMG.exists():
        st.image(
            str(ARQUIVO_CARDAPIO_IMG),
            caption="Cardápio JMD Assados",
            use_container_width=True
        )
    else:
        st.warning(
            "Imagem do cardápio não encontrada. "
            "Verifique se o arquivo `jmd.png` está no GitHub, na mesma pasta do app."
        )


def main():
    aplicar_css()
    inicializar_estado()

    st.title("🍗 Chatbot WhatsApp - Loja de Assados")
    st.caption("Protótipo em formato de chat para apresentação no Streamlit.")

    if not Path(ARQUIVO_CONFIG).exists():
        st.error(
            f"Arquivo `{ARQUIVO_CONFIG}` não encontrado. "
            "Coloque a planilha na mesma pasta do app antes de rodar."
        )

        st.write("Arquivos encontrados nesta pasta:")
        try:
            st.write([p.name for p in BASE_DIR.iterdir()])
        except Exception:
            pass

        st.stop()

    if not carregar_cardapio():
        st.error("Nenhum item disponível foi encontrado na aba Cardapio.")
        st.stop()

    exibir_sidebar()

    col_chat, col_final = st.columns([1.2, 0.8])

    with col_chat:
        for message in st.session_state.messages:
            with st.chat_message(message["role"]):
                conteudo = message["content"]
                conteudo = conteudo.replace("R$", "R\\$")
                conteudo = conteudo.replace("\n", "  \n")
                st.markdown(conteudo)

        prompt = st.chat_input(
            placeholder_chat(),
            key="campo_chat_principal"
        )

        rolar_e_focar_chat()

        if prompt:
            add_user(prompt)
            resposta = processar_mensagem(prompt)
            add_bot(resposta)
            st.rerun()

    with col_final:
        exibir_cardapio_imagem()
        exibir_resumo_final_na_tela()


if __name__ == "__main__":
    main()
